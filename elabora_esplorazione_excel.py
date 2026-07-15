"""
Elabora un UNICO file (Excel o CSV) contenente i dati di più osservazioni
(sessioni), distinte dalla colonna 'Observation id'. La fase
(training / test) viene dedotta dal testo di 'Observation id' stesso
(es. 'c3m1_test' -> test, 'c3m1_training' -> training).

Il file può contenere una o più sessioni: ogni valore univoco di
'Observation id' viene trattato come una sessione indipendente, con la
propria somma cumulata di esplorazione (che riparte da zero per ciascuna
sessione).

Output (in the OUTPUT_FOLDER):
1. 'data_complete_formulas.xlsx'
   -> all data ordered by Observation id and Start (s), with added columns
      'Cumulative (s)' and 'Explore_first10s' AS EXCEL FORMULAS
      (automatically recalculated when the file is opened in Excel). The
      cumulative sum restarts from zero for each new 'Observation id'.
2. 'first10s_events.xlsx'
   -> only the rows that belong to the first SOGLIA_SECONDI seconds of
      cumulative exploration for each session.
3. 'session_summary.xlsx'
   -> one row per 'Observation id', with Phase (training/test),
      Discrimination Index, latency to first exploration event (overall
      and per type), time needed to reach SOGLIA_SECONDI seconds of
      exploration, and for each exploration behavior type found in the file
      (e.g. Explore_F, Explore_F2, Explore_N): total, mean bout duration,
      and number of events. These metrics are calculated only on the first
      SOGLIA_SECONDI seconds of exploration (except per-type latencies,
      which are computed over the full session).
      The summary also includes, for each time bin among N_BIN bins of
      DURATA_BIN_MINUTI minutes (based on session Start (s), not limited
      to the first 10s): total exploration time and the bin Discrimination
      Index (N vs F for test, F2 vs F for training).

Usage:
    python elabora_esplorazione_excel.py

Requirements: pip install pandas openpyxl
"""

import os
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ------------------- PARAMETRI MODIFICABILI -------------------
INPUT_FILE = "full data.xlsx"   # single file with ALL data (.xlsx or .csv)
OUTPUT_FOLDER = "output"
OUTPUT_FORMULAS = "data_complete_formulas.xlsx"
OUTPUT_FIRST10S = "first10s_events.xlsx"
OUTPUT_SUMMARY = "session_summary.xlsx"
SOGLIA_SECONDI = 10
PREFISSO_BEHAVIOR = "Explore"       # prefix for exploration behavior types

OBS_ID_COL_NAME = "Observation id"
BEHAVIOR_COL_NAME = "Behavior"
START_COL_NAME = "Start (s)"
STOP_COL_NAME = "Stop (s)"
DURATION_COL_NAME = "Duration (s)"

DURATA_BIN_MINUTI = 2   # ampiezza di ciascun time bin, in minuti
N_BIN = 5                # numero di bin (5 x 2min = primi 10 minuti di sessione)
# ----------------------------------------------------------------


def carica_dati(path):
    """Legge il file di input, sia esso .xlsx o .csv."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return pd.read_csv(path)
    return pd.read_excel(path)


OBS_ID_NUM_RE = re.compile(r"c(\d+)", re.IGNORECASE)


def estrai_fase(observation_id):
    """Deduce 'training' o 'test' dal testo di Observation id.
    Modifica qui se i tuoi nomi usano parole diverse (es. 'baseline')."""
    testo = str(observation_id).lower()
    if "train" in testo:
        return "training"
    if "test" in testo:
        return "test"
    return "non specified"


def estrai_numero_observation_id(observation_id):
    testo = str(observation_id)
    match = OBS_ID_NUM_RE.search(testo)
    return int(match.group(1)) if match else float("inf")


def get_first10s_subset(df_sessione):
    """Dato il sotto-dataframe di UNA sessione (un solo Observation id),
    restituisce solo le righe di esplorazione entro i primi SOGLIA_SECONDI
    secondi di esplorazione cumulata, con la colonna 'Cumulative (s)'."""
    is_explore = df_sessione[BEHAVIOR_COL_NAME].astype(str).str.startswith(PREFISSO_BEHAVIOR)
    explore_df = df_sessione[is_explore].sort_values(START_COL_NAME).copy()
    explore_df["Cumulative (s)"] = explore_df[DURATION_COL_NAME].cumsum()
    prima_soglia = (explore_df["Cumulative (s)"] - explore_df[DURATION_COL_NAME]) < SOGLIA_SECONDI
    return explore_df.loc[prima_soglia]


def calcola_metriche_bin(df_sessione, fase):
    """Divide l'intera sessione in bin temporali da DURATA_BIN_MINUTI minuti
    (in base allo 'Start (s)' di ciascun evento) e calcola, per ogni bin:
    - il tempo di esplorazione TOTALE (somma di tutti i tipi Explore*)
    - il Discrimination Index del bin: N vs F per le sessioni test,
      F2 vs F per le sessioni training (stessa convenzione usata altrove)
    """
    beh = df_sessione[BEHAVIOR_COL_NAME].astype(str)
    is_explore = beh.str.startswith(PREFISSO_BEHAVIOR)
    dur = df_sessione[DURATION_COL_NAME]
    start = df_sessione[START_COL_NAME]

    bin_sec = DURATA_BIN_MINUTI * 60
    colonne_bin = {}

    for i in range(N_BIN):
        inizio = i * bin_sec
        fine = (i + 1) * bin_sec
        min_inizio = i * DURATA_BIN_MINUTI
        min_fine = (i + 1) * DURATA_BIN_MINUTI
        etichetta = f"{min_inizio}-{min_fine}min"

        in_bin = is_explore & (start >= inizio) & (start < fine)
        tempo_totale_bin = dur[in_bin].sum()

        somma_N_bin = dur[in_bin & (beh == "Explore_N")].sum()
        somma_F_bin = dur[in_bin & (beh == "Explore_F")].sum()
        somma_F2_bin = dur[in_bin & (beh == "Explore_F2")].sum()

        if fase == "test":
            denom_bin = somma_N_bin + somma_F_bin
            DI_bin = (somma_N_bin - somma_F_bin) / denom_bin if denom_bin != 0 else float("nan")
        elif fase == "training":
            denom_bin = somma_F2_bin + somma_F_bin
            DI_bin = (somma_F2_bin - somma_F_bin) / denom_bin if denom_bin != 0 else float("nan")
        else:
            DI_bin = float("nan")

        colonne_bin[f"Discrimination Index bin {i+1} ({etichetta})"] = DI_bin
        colonne_bin[f"Total exploration time bin {i+1} ({etichetta}) (s)"] = tempo_totale_bin

    return colonne_bin


def riassumi_sessione(observation_id, df_sessione, behavior_types):
    """Calcola le metriche per una sessione, solo sui primi SOGLIA_SECONDI secondi.
    behavior_types: lista di TUTTI i tipi di comportamento di esplorazione trovati
    nell'intero file (es. ['Explore_F', 'Explore_F2', 'Explore_N']), così che il
    riepilogo abbia sempre le stesse colonne per ogni sessione."""
    subset = get_first10s_subset(df_sessione)
    fase = estrai_fase(observation_id)

    beh = subset[BEHAVIOR_COL_NAME].astype(str)
    dur = subset[DURATION_COL_NAME]
    start = subset[START_COL_NAME]
    stop = subset[STOP_COL_NAME]

    # Statistiche per ogni tipo di comportamento (usate sia per le colonne
    # individuali sia per calcolare i due Discrimination Index)
    somma_per_tipo = {}
    per_tipo_colonne = {}
    beh_sessione = df_sessione[BEHAVIOR_COL_NAME].astype(str)
    for tipo in behavior_types:
        is_tipo = beh == tipo
        n_eventi = int(is_tipo.sum())
        mean_bout = dur[is_tipo].mean() if is_tipo.any() else float("nan")
        somma_tipo = dur[is_tipo].sum()
        somma_per_tipo[tipo] = somma_tipo

        # Latenza al primo evento di QUESTO tipo: calcolata sull'INTERA
        # sessione (non solo sul subset primi 10s), perché il primo evento
        # di un tipo specifico potrebbe avvenire dopo la soglia dei 10s
        # cumulati di esplorazione totale.
        is_tipo_sessione = beh_sessione == tipo
        latenza_tipo = (
            df_sessione.loc[is_tipo_sessione, START_COL_NAME].min()
            if is_tipo_sessione.any() else float("nan")
        )

        per_tipo_colonne[f"Latency first {tipo} event (s)"] = latenza_tipo
        per_tipo_colonne[f"Total {tipo} first 10s (s)"] = somma_tipo
        per_tipo_colonne[f"Mean bout duration {tipo} first 10s (s)"] = mean_bout
        per_tipo_colonne[f"N events {tipo} first 10s"] = n_eventi

    somma_N = somma_per_tipo.get("Explore_N", 0.0)
    somma_F = somma_per_tipo.get("Explore_F", 0.0)
    somma_F2 = somma_per_tipo.get("Explore_F2", 0.0)

    # DI for first 10s: test uses Explore_N vs Explore_F, training uses Explore_F2 vs Explore_F
    if fase == "test":
        denom = somma_N + somma_F
        DI_first10s = (somma_N - somma_F) / denom if denom != 0 else float("nan")
    elif fase == "training":
        denom_training = somma_F2 + somma_F
        DI_first10s = (somma_F2 - somma_F) / denom_training if denom_training != 0 else float("nan")
    else:
        DI_first10s = float("nan")

    # Overall 10-minute window on session time (0-10 min)
    mask_10min = (
        beh_sessione.str.startswith(PREFISSO_BEHAVIOR)
        & (df_sessione[START_COL_NAME] < 10 * 60)
    )
    sum_explore_N_10min = df_sessione.loc[mask_10min & (beh_sessione == "Explore_N"), DURATION_COL_NAME].sum()
    sum_explore_F_10min = df_sessione.loc[mask_10min & (beh_sessione == "Explore_F"), DURATION_COL_NAME].sum()
    sum_explore_F2_10min = df_sessione.loc[mask_10min & (beh_sessione == "Explore_F2"), DURATION_COL_NAME].sum()
    total_exploration_10min = df_sessione.loc[mask_10min, DURATION_COL_NAME].sum()

    # DI for the full 10-minute window, using the total durations for these
    # Explore_* events over the entire 10-minute period.
    if fase == "test":
        denom_10min = sum_explore_N_10min + sum_explore_F_10min
        DI_10min = (sum_explore_N_10min - sum_explore_F_10min) / denom_10min if denom_10min != 0 else float("nan")
    elif fase == "training":
        denom_10min = sum_explore_F2_10min + sum_explore_F_10min
        DI_10min = (sum_explore_F2_10min - sum_explore_F_10min) / denom_10min if denom_10min != 0 else float("nan")
    else:
        DI_10min = float("nan")

    latenza = start.min() if len(start) else float("nan")

    # Tempo di sessione necessario a raggiungere i 10s di esplorazione cumulata
    tempo_a_soglia = stop.iloc[-1] if len(stop) else float("nan")

    # Nelle sessioni 'training' non ha senso Explore_N (niente oggetto novel):
    # svuotiamo le sue colonne e il DI classico che lo usa.
    if fase == "training":
        DI = float("nan")
        for chiave in list(per_tipo_colonne):
            if "Explore_N" in chiave:
                per_tipo_colonne[chiave] = float("nan")

    # Nelle sessioni 'test' non ha senso Explore_F2 (niente secondo familiare):
    # svuotiamo le sue colonne (il DI training è già NaN per queste sessioni).
    if fase == "test":
        for chiave in list(per_tipo_colonne):
            if "Explore_F2" in chiave:
                per_tipo_colonne[chiave] = float("nan")

    risultato = {
        "Observation id": observation_id,
        "Phase": fase,
        "Discrimination Index (first 10s)": DI_first10s,
    }

    bin_metrics = calcola_metriche_bin(df_sessione, fase)
    di_bin_metrics = {k: v for k, v in bin_metrics.items() if k.startswith("Discrimination Index bin")}
    time_bin_metrics = {k: v for k, v in bin_metrics.items() if not k.startswith("Discrimination Index bin")}

    risultato.update(di_bin_metrics)
    risultato.update({
        "Discrimination Index all 10 minutes": DI_10min,
        "Total exploration time all 10 minutes (s)": total_exploration_10min,
    })
    risultato.update({
        "Latency first exploration event (s)": latenza,
        f"Time to reach {SOGLIA_SECONDI}s of exploration (s)": tempo_a_soglia,
    })
    risultato.update(per_tipo_colonne)
    risultato.update(time_bin_metrics)
    return risultato


def salva_file_con_formule(df_ordinato, output_folder):
    """Scrive un xlsx con le colonne 'Cumulative (s)' e 'Explore_first10s'
    come FORMULE Excel vere. La somma cumulata riparte da zero ogni volta
    che cambia 'Observation id' (grazie al criterio SUMIFS sull'Observation id)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dati"

    for row in dataframe_to_rows(df_ordinato, index=False, header=True):
        ws.append(row)

    header = [cell.value for cell in ws[1]]
    col_obsid = header.index(OBS_ID_COL_NAME) + 1
    col_behavior = header.index(BEHAVIOR_COL_NAME) + 1
    col_duration = header.index(DURATION_COL_NAME) + 1

    n_righe = ws.max_row
    col_cum = ws.max_column + 1
    col_flag = ws.max_column + 2
    ws.cell(row=1, column=col_cum, value="Cumulative (s)")
    ws.cell(row=1, column=col_flag, value="Explore_first10s")

    l_obs = get_column_letter(col_obsid)
    l_beh = get_column_letter(col_behavior)
    l_dur = get_column_letter(col_duration)
    l_cum = get_column_letter(col_cum)

    for r in range(2, n_righe + 1):
        formula_cum = (
            f'=SUMIFS(${l_dur}$2:{l_dur}{r},'
            f'${l_obs}$2:{l_obs}{r},{l_obs}{r},'
            f'${l_beh}$2:{l_beh}{r},"{PREFISSO_BEHAVIOR}*")'
        )
        ws.cell(row=r, column=col_cum, value=formula_cum)

        formula_flag = (
            f'=IF(AND(LEFT({l_beh}{r},7)="{PREFISSO_BEHAVIOR}",'
            f'({l_cum}{r}-IF(LEFT({l_beh}{r},7)="{PREFISSO_BEHAVIOR}",'
            f'{l_dur}{r},0))<{SOGLIA_SECONDI}),"SI","NO")'
        )
        ws.cell(row=r, column=col_flag, value=formula_flag)

    os.makedirs(output_folder, exist_ok=True)
    path_out = os.path.join(output_folder, OUTPUT_FORMULAS)
    wb.save(path_out)
    print(f"[OK] Formulas added (per session) -> {path_out}")
    print("     (open the file in Excel: the formulas are calculated automatically)")


def main():
    if not os.path.exists(INPUT_FILE):
        print(f"File non trovato: '{INPUT_FILE}'. Modifica INPUT_FILE con il percorso corretto.")
        return

    df = carica_dati(INPUT_FILE)
    if OBS_ID_COL_NAME not in df.columns:
        print(f"Colonna '{OBS_ID_COL_NAME}' non trovata nel file.")
        return

    df_ordinato = df.copy()
    df_ordinato["Phase"] = df_ordinato[OBS_ID_COL_NAME].apply(estrai_fase)
    df_ordinato["_phase_order"] = df_ordinato["Phase"].map({"training": 0, "test": 1, "non specified": 2}).fillna(99)
    df_ordinato["_obs_number"] = df_ordinato[OBS_ID_COL_NAME].apply(estrai_numero_observation_id)
    df_ordinato = df_ordinato.sort_values(
        ["_phase_order", "_obs_number", OBS_ID_COL_NAME, START_COL_NAME]
    ).drop(columns=["Phase", "_phase_order", "_obs_number"]).reset_index(drop=True)

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    # Tutti i tipi di comportamento di esplorazione presenti nel file
    # (es. Explore_F, Explore_F2, Explore_N...), in ordine alfabetico
    behavior_types = sorted(
        df_ordinato.loc[
            df_ordinato[BEHAVIOR_COL_NAME].astype(str).str.startswith(PREFISSO_BEHAVIOR),
            BEHAVIOR_COL_NAME
        ].astype(str).unique()
    )
    print(f"Tipi di esplorazione trovati: {behavior_types}")

    # 1) File con le formule (cumulata + flag), per sessione
    salva_file_con_formule(df_ordinato, OUTPUT_FOLDER)

    # 2) File con solo gli eventi entro i primi 10s, di tutte le sessioni
    tutti_primi10s = []
    riepilogo = []
    for obs_id, gruppo in df_ordinato.groupby(OBS_ID_COL_NAME, sort=False):
        subset = get_first10s_subset(gruppo)
        subset = subset[[OBS_ID_COL_NAME, BEHAVIOR_COL_NAME, START_COL_NAME,
                          STOP_COL_NAME, DURATION_COL_NAME, "Cumulative (s)"]].copy()
        subset.insert(1, "Phase", estrai_fase(obs_id))
        tutti_primi10s.append(subset)
        riepilogo.append(riassumi_sessione(obs_id, gruppo, behavior_types))

    df_primi10s = pd.concat(tutti_primi10s, ignore_index=True)
    df_primi10s["_phase_order"] = df_primi10s["Phase"].map({"training": 0, "test": 1, "non specified": 2}).fillna(99)
    df_primi10s["_obs_number"] = df_primi10s[OBS_ID_COL_NAME].apply(estrai_numero_observation_id)
    df_primi10s = df_primi10s.sort_values(
        ["_phase_order", "_obs_number", OBS_ID_COL_NAME, START_COL_NAME]
    ).drop(columns=["_phase_order", "_obs_number"]).reset_index(drop=True)
    path_primi10s = os.path.join(OUTPUT_FOLDER, OUTPUT_FIRST10S)
    df_primi10s.to_excel(path_primi10s, index=False)
    print(f"[OK] Events within {SOGLIA_SECONDI}s for all sessions -> {path_primi10s}")

    # 3) Riepilogo con una riga per sessione (Observation id)
    df_riepilogo = pd.DataFrame(riepilogo)
    df_riepilogo["_phase_order"] = df_riepilogo["Phase"].map({"training": 0, "test": 1, "non specified": 2}).fillna(99)
    df_riepilogo["_obs_number"] = df_riepilogo["Observation id"].apply(estrai_numero_observation_id)
    df_riepilogo = df_riepilogo.sort_values(
        ["_phase_order", "_obs_number", "Observation id"]
    ).drop(columns=["_phase_order", "_obs_number"]).reset_index(drop=True)
    path_riepilogo = os.path.join(OUTPUT_FOLDER, OUTPUT_SUMMARY)
    df_riepilogo.to_excel(path_riepilogo, index=False)
    print(f"[OK] Session summary ({len(df_riepilogo)} sessions) -> {path_riepilogo}")


if __name__ == "__main__":
    main()
